import openpyxl

# ścieżka do pliku tekstowego z danymi
txt_path = "Wyniki BST - balancer.txt"

# lista do przechowywania danych
data = []

# otwieranie pliku tekstowego i odczytywanie danych
with open(txt_path, "r") as f:
    for line in f:
        if line.strip():  # sprawdza, czy linia nie jest pusta
            parts = line.split()
            data.append({"N": int(parts[0]), "czas": float(parts[1])})

# tworzenie słownika ze średnimi czasami dla poszczególnych N
averages = {}
for d in data:
    n = d["N"]
    czas = d["czas"]
    if n in averages:
        averages[n].append(czas)
    else:
        averages[n] = [czas]

# obliczanie średnich czasów
for n in averages:
    srednia = sum(averages[n]) / len(averages[n])
    averages[n] = srednia

# zapisywanie wyników do pliku Excel
excel_path = "Wyniki BST - balancer.xlsx"

wb = openpyxl.Workbook()
ws = wb.active

# nagłówki kolumn
ws.cell(row=1, column=1).value = "N"
ws.cell(row=1, column=2).value = "Średni czas"

# wprowadzanie danych
row = 2
for n in averages:
    ws.cell(row=row, column=1).value = n
    ws.cell(row=row, column=2).value = averages[n]
    row += 1

# zapisywanie pliku Excel
wb.save(excel_path)
